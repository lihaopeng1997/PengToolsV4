# -*- coding: utf-8 -*-
import copy
import datetime
import os
import re
from urllib.parse import unquote, urlsplit

from openpyxl import load_workbook


RELEASE_SVN_URL = 'svn://10.128.23.145:13690/YDPIC_SharingCenter/17.集成测试/02-生产环境发版任务清单'
RELEASE_WORKBOOK_NAME = '英大财险-共享中心-生产任务发版清单.xlsx'
RELEASE_HEADERS = (
    '系统名', '分支名称', '前端：全部', '任务编号', '任务内容', '升级公告', '编码人',
    '代码是否已合并', '负责人', '计划升级日期', '验证组测试是否通过', '备注',
    '系统操作改变', '岗位菜单', '历史数据', '定时任务', '开关新增', '数据质量校验',
    '平台/三方交互', '相关文档', '是否需要验证工号', '验证机构', '需要权限',
)


def branch_name_from_svn(url):
    value = str(url or '').strip().rstrip('/\\')
    if not value:
        return ''
    if '://' in value:
        value = urlsplit(value).path.rstrip('/').rsplit('/', 1)[-1]
    else:
        value = re.split(r'[/\\]', value)[-1]
    return unquote(value)


def requirement_date(requirement):
    return str(requirement.get('actual_online_date') or requirement.get('planned_online_date') or '').strip()[:10]


def rank_requirements(requirements, upgrade_date):
    target = upgrade_date.strftime('%Y-%m-%d') if hasattr(upgrade_date, 'strftime') else str(upgrade_date)
    month = target[:7]
    def score(item):
        date_value = requirement_date(item)
        if date_value == target:
            rank = 0
        elif date_value.startswith(month):
            rank = 1
        elif not date_value:
            rank = 3
        else:
            rank = 2
        return rank, date_value or '9999-99-99', item.get('record_kind', '需求'), item.get('code', '')
    return sorted(requirements, key=score)


def release_row_from_requirement(requirement, upgrade_date, values=None):
    values = values or {}
    target_date = upgrade_date.date() if isinstance(upgrade_date, datetime.datetime) else upgrade_date
    if isinstance(target_date, str):
        target_date = datetime.datetime.strptime(target_date.replace('-', ''), '%Y%m%d').date()
    branch = values.get('分支名称') or branch_name_from_svn(requirement.get('svn_url'))
    task_code = values.get('任务编号') or requirement.get('code') or requirement.get('title', '')
    content = values.get('任务内容') or requirement.get('description') or requirement.get('title', '')
    announcement = values.get('升级公告') or requirement.get('title') or content
    result = {
        '系统名': values.get('系统名') or requirement.get('system') or '车险承保中心',
        '分支名称': branch,
        '前端：全部': values.get('前端：全部') or '后端：全部',
        '任务编号': task_code,
        '任务内容': content,
        '升级公告': announcement,
        '编码人': values.get('编码人') or requirement.get('owner') or '',
        '代码是否已合并': values.get('代码是否已合并') or '否',
        '负责人': values.get('负责人') or requirement.get('owner') or '',
        '计划升级日期': target_date,
        '验证组测试是否通过': values.get('验证组测试是否通过') or '否',
        '备注': values.get('备注') or '',
        '系统操作改变': values.get('系统操作改变') or '否',
        '岗位菜单': values.get('岗位菜单') or '否',
        '历史数据': values.get('历史数据') or '否',
        '定时任务': values.get('定时任务') or '否',
        '开关新增': values.get('开关新增') or '否',
        '数据质量校验': values.get('数据质量校验') or '否',
        '平台/三方交互': values.get('平台/三方交互') or ('是' if requirement.get('needs_peripheral_upgrade') else '否'),
        '相关文档': values.get('相关文档') or ('是' if requirement.get('needs_interface_update') else '否'),
        '是否需要验证工号': values.get('是否需要验证工号') or '否',
        '验证机构': values.get('验证机构') or '全辖',
        '需要权限': values.get('需要权限') or '',
    }
    return result


def _dated_sheets(workbook):
    return [sheet for sheet in workbook.worksheets if re.fullmatch(r'20\d{6}', sheet.title)]


def _last_data_row(sheet):
    for row in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row, column).value not in (None, '') for column in range(1, len(RELEASE_HEADERS) + 1)):
            return row
    return 1


def _copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, len(RELEASE_HEADERS) + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)


def _new_date_sheet(workbook, sheet_name):
    dated = _dated_sheets(workbook)
    if not dated:
        raise ValueError('发版清单中没有可复制的日期工作表')
    candidates = [sheet for sheet in dated if sheet.title < sheet_name]
    template = candidates[-1] if candidates else dated[-1]
    sheet = workbook.copy_worksheet(template)
    sheet.title = sheet_name
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 2:
            sheet.unmerge_cells(str(merged))
    for row in range(2, max(sheet.max_row, 50) + 1):
        if row > sheet.max_row:
            _copy_row_style(sheet, max(2, sheet.max_row), row)
        for column in range(1, len(RELEASE_HEADERS) + 1):
            sheet.cell(row, column).value = None
    return sheet


def update_release_workbook(template_path, output_path, upgrade_date, rows):
    sheet_name = upgrade_date.strftime('%Y%m%d') if hasattr(upgrade_date, 'strftime') else str(upgrade_date).replace('-', '')
    if not re.fullmatch(r'20\d{6}', sheet_name):
        raise ValueError('升级日期必须为 yyyyMMdd')
    if not rows:
        raise ValueError('请至少选择一个需求或 BUG')
    workbook = load_workbook(template_path)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else _new_date_sheet(workbook, sheet_name)
        headers = [sheet.cell(1, column).value for column in range(1, len(RELEASE_HEADERS) + 1)]
        if tuple(headers) != RELEASE_HEADERS:
            raise ValueError(f'{sheet.title} 的 23 列表头与生产清单模板不一致')
        start_row = _last_data_row(sheet) + 1
        for offset, row_values in enumerate(rows):
            target_row = start_row + offset
            if target_row > sheet.max_row:
                _copy_row_style(sheet, max(2, target_row - 1), target_row)
            for column, header in enumerate(RELEASE_HEADERS, 1):
                cell = sheet.cell(target_row, column)
                cell.value = row_values.get(header)
                if header == '计划升级日期':
                    cell.number_format = 'yyyy-mm-dd'
            if not sheet.row_dimensions[target_row].height:
                sheet.row_dimensions[target_row].height = 27
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()
    return {'path': output_path, 'sheet_name': sheet_name, 'start_row': start_row, 'end_row': start_row + len(rows) - 1}
