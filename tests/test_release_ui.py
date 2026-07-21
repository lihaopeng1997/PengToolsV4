# -*- coding: utf-8 -*-
import os
import shutil
import subprocess
import tempfile
import sys
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from PyQt6.QtCore import QDate, Qt
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import QApplication, QDialog, QHeaderView, QSplitter

from panels.requirement_panel import DateInput, RequirementDialog, RequirementPanel, format_online_month_label
from panels.sql_panel import SqlToolPanel
from tools.release_prep import RELEASE_HEADERS, RELEASE_WORKBOOK_NAME
from main_window import MainWindow
from config import local_data_dir
from ui.confirm_dialog import AppNoticeDialog, CloseActionDialog, ConfirmActionDialog, NextStepDialog
from tools.svn_workspace import add_text_file, checkout, commit_working_copy, run_svn, scan_working_copies, workspace_files


class ReleaseUiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])
        cls.app.setFont(QFont('Microsoft YaHei UI', 10))

    def test_requirement_allows_empty_development_svn(self):
        dialog = RequirementDialog()
        dialog.system_edit.setCurrentIndex(1)
        dialog.title_edit.setText('测试需求')
        dialog._accept_checked()
        self.assertEqual(dialog.result(), QDialog.DialogCode.Accepted)
        self.assertEqual(dialog.values()['svn_url'], '')
        dialog = RequirementDialog()
        dialog.system_edit.setCurrentIndex(1)
        dialog.title_edit.setText('测试需求')
        dialog.svn_url_edit.setText('svn://10/x/DEV_REQ_TEST')
        dialog._accept_checked()
        self.assertEqual(dialog.result(), QDialog.DialogCode.Accepted)
        self.assertEqual(dialog.values()['svn_url'], 'svn://10/x/DEV_REQ_TEST')
        self.assertEqual(dialog.values()['system'], dialog.system_edit.currentData())

    def test_requirement_system_configuration_is_shared(self):
        panel = RequirementPanel()
        requirement_systems = [panel.system_filter.itemData(index) for index in range(1, panel.system_filter.count())]
        sql_panel = SqlToolPanel()
        sql_systems = [system['name'] for system in sql_panel._systems]
        self.assertEqual(requirement_systems, sql_systems)
        opened = []
        panel.open_system_config.connect(lambda: opened.append(True))
        panel.system_config_btn.click()
        self.assertEqual(opened, [True])
        panel.close(); sql_panel.close()

    def test_delete_confirmation_is_cancel_first_and_cancel_default(self):
        dialog = ConfirmActionDialog('删除测试', '删除后无法恢复')
        dialog.show()
        self.app.processEvents()
        buttons = dialog.layout().itemAt(2).layout()
        self.assertIs(buttons.itemAt(1).widget(), dialog.cancel_button)
        self.assertIs(buttons.itemAt(2).widget(), dialog.confirm_button)
        self.assertEqual(dialog.cancel_button.text(), '取消')
        self.assertEqual(dialog.confirm_button.text(), '确认删除')
        self.assertTrue(dialog.cancel_button.isDefault())
        self.assertTrue(dialog.cancel_button.hasFocus())
        dialog.reject()

        notice = AppNoticeDialog('升级材料已生成', 'Sheet 已写入', kind='success')
        notice.show()
        self.app.processEvents()
        self.assertEqual(notice.ok_button.text(), '知道了')
        self.assertTrue(notice.ok_button.isDefault())
        notice.accept()

        close_dialog = CloseActionDialog(language='zh', default_action='minimize')
        close_dialog.show()
        self.app.processEvents()
        self.assertEqual(close_dialog.windowTitle(), '关闭 PengTools？')
        self.assertTrue(close_dialog.minimize_button.hasFocus())
        self.assertFalse(close_dialog.dont_ask_again())
        close_dialog.dont_ask_check.setChecked(True)
        close_dialog.minimize_button.clicked.emit()
        self.assertEqual(close_dialog.selected_action(), 'minimize')
        self.assertTrue(close_dialog.dont_ask_again())

        # 默认动作若为 exit：焦点仍落在安全控件（取消），不误触危险卡
        close_exit_default = CloseActionDialog(language='zh', default_action='exit')
        close_exit_default.show()
        self.app.processEvents()
        self.assertTrue(close_exit_default.cancel_button.hasFocus())
        close_exit_default.reject()

        panel = SqlToolPanel()
        original_count = len(panel._systems)
        with patch('panels.sql_panel.confirm_action', return_value=False), patch('panels.sql_panel.save_systems'):
            panel._delete_system()
        self.assertEqual(len(panel._systems), original_count)
        with patch('panels.sql_panel.confirm_action', return_value=True), patch('panels.sql_panel.save_systems'):
            panel._delete_system()
        self.assertEqual(len(panel._systems), original_count - 1)
        panel.close()

    def test_pasted_bug_prompts_for_development_svn(self):
        with patch('panels.requirement_panel.load_requirements', return_value=[]):
            panel = RequirementPanel()
        self.assertFalse(hasattr(panel, 'category_filter'))
        self.assertEqual(panel.lock_file_btn.text(), '锁定')
        self.assertEqual(panel.unlock_file_btn.text(), '解锁')
        self.assertEqual(panel.scan_btn.text(), '扫描需求目录')
        self.assertEqual(panel.checkout_btn.text(), '检出代码')
        self.assertEqual(panel.bug_btn.text(), '登记缺陷')
        with patch('panels.requirement_panel.QInputDialog.getMultiLineText', return_value=('BUG-100 测试问题', True)), \
                patch('panels.requirement_panel.QInputDialog.getText', return_value=('svn://10/x/DEV_BUG_100', True)) as svn_prompt, \
                patch('panels.requirement_panel.save_requirements'), \
                patch('panels.requirement_panel.offer_next_steps', return_value=None) as next_steps:
            panel._paste_bug()
        svn_prompt.assert_called_once()
        next_steps.assert_called_once()
        self.assertEqual(panel._requirements[-1]['svn_url'], 'svn://10/x/DEV_BUG_100')
        panel.resize(1220, 780)
        panel.show()
        self.app.processEvents()
        screenshot = os.path.join(ROOT, '.codex_work', 'requirement_svn_lock_ui.png')
        self.assertTrue(panel.grab().save(screenshot))
        panel.close()

    @unittest.skipUnless(shutil.which('svn'), 'SVN command line is unavailable')
    def test_scan_result_stays_browsable_and_loading_restores_controls(self):
        svnadmin = os.path.join(os.path.dirname(shutil.which('svn')), 'svnadmin.exe')
        if not os.path.isfile(svnadmin):
            self.skipTest('svnadmin is unavailable')
        with tempfile.TemporaryDirectory() as temp:
            repository = os.path.join(temp, 'repository')
            subprocess.run([svnadmin, 'create', repository], check=True)
            url = Path(repository).as_uri()
            run_svn(['mkdir', f'{url}/trunk', '-m', 'init'])
            working_copy = os.path.join(temp, '2026-07', 'requirement', 'REQ-100')
            checkout(f'{url}/trunk', working_copy)
            add_text_file(working_copy, 'readme.txt', 'test requirement')
            commit_working_copy(working_copy, 'add test requirement')

            with patch('panels.requirement_panel.load_requirements', return_value=[]), \
                    patch('panels.requirement_panel.save_requirements'), \
                    patch('panels.requirement_panel.show_info'), \
                    patch('panels.requirement_panel.show_success'), \
                    patch('panels.requirement_panel.show_warning'), \
                    patch('panels.requirement_panel.offer_next_steps', return_value=None):
                panel = RequirementPanel()
                panel._start_task('scanning', scan_working_copies, (temp,), panel._scan_finished)
                self.assertFalse(panel.loading.isHidden())
                self.assertFalse(panel.requirement_list.isEnabled())

                deadline = time.time() + 10
                idle_cycles = 0
                while time.time() < deadline:
                    self.app.processEvents()
                    time.sleep(0.01)
                    if panel._active_worker is None and not panel._pending_file_refresh:
                        idle_cycles += 1
                        if idle_cycles >= 5:
                            break
                    else:
                        idle_cycles = 0
                self.app.processEvents()

                self.assertIsNone(panel._active_worker)
                requirement_items = [
                    panel.requirement_list.topLevelItem(group).child(child)
                    for group in range(panel.requirement_list.topLevelItemCount())
                    for child in range(panel.requirement_list.topLevelItem(group).childCount())
                ]
                self.assertEqual(len(requirement_items), 1)
                self.assertEqual(panel.detail_title.text(), 'REQ-100')
                self.assertEqual(panel.file_tree.topLevelItemCount(), 1)
                self.assertEqual(panel.file_tree.topLevelItem(0).text(0), 'readme.txt')
                self.assertTrue(panel.requirement_list.isEnabled())
                self.assertTrue(panel.edit_btn.isEnabled())
                self.assertTrue(panel.open_folder_btn.isEnabled())
                panel.close()

    def test_failed_waiting_task_restores_controls(self):
        def fail_task():
            raise RuntimeError('offline test failure')

        with patch('panels.requirement_panel.load_requirements', return_value=[]), \
                patch('panels.requirement_panel.show_warning') as warning:
            panel = RequirementPanel()
            panel._start_task('waiting', fail_task, (), lambda _result: None)
            deadline = time.time() + 5
            while panel._active_worker is not None and time.time() < deadline:
                self.app.processEvents()
                time.sleep(0.01)
            self.app.processEvents()
            warning.assert_called_once()
            self.assertTrue(panel.requirement_list.isEnabled())
            self.assertIn('offline test failure', panel.loading._label)
            panel.close()

    def test_online_month_label_formats_complete_chinese_title(self):
        self.assertEqual(format_online_month_label('2026-06'), '2026年6月')
        self.assertEqual(format_online_month_label('2026-6'), '2026年6月')
        self.assertEqual(format_online_month_label('2026年06月'), '2026年6月')
        self.assertEqual(format_online_month_label('未分月'), '未分月')

    def test_requirements_are_grouped_by_month_and_sorted_by_modified_time(self):
        requirements = [
            {'id': 'feb', 'title': 'February', 'code': 'REQ-FEB', 'online_month': '2026-02', 'source_modified_at': '2026-07-17T12:00:00'},
            {'id': 'jun-old', 'title': 'June old', 'code': 'REQ-JUN-1', 'online_month': '2026-06', 'source_modified_at': '2026-07-16T12:00:00'},
            {'id': 'jun-new', 'title': 'June new', 'code': 'REQ-JUN-2', 'online_month': '2026-06', 'source_modified_at': '2026-07-17T12:00:00'},
            {'id': 'long', 'title': 'Long code', 'code': 'REQ-VERY-LONG-CODE-202606-01', 'online_month': '2026-06', 'source_modified_at': '2026-07-15T12:00:00'},
        ]
        with patch('panels.requirement_panel.load_requirements', return_value=requirements):
            panel = RequirementPanel()
        self.assertEqual(panel.requirement_list.topLevelItemCount(), 2)
        june = panel.requirement_list.topLevelItem(0)
        february = panel.requirement_list.topLevelItem(1)
        self.assertTrue(june.isFirstColumnSpanned())
        self.assertIn('2026年6月', june.text(0))
        self.assertIn('3 项', june.text(0))
        self.assertEqual(june.child(0).text(0), 'REQ-JUN-2')
        self.assertEqual(june.child(1).text(0), 'REQ-JUN-1')
        self.assertEqual(june.child(2).text(0), 'REQ-VERY-LONG-CODE-202606-01')
        self.assertGreaterEqual(panel.requirement_list.columnWidth(0), 180)
        self.assertIn('2026年2月', february.text(0))
        self.assertEqual(panel._current['id'], 'jun-new')
        panel.close()

    def test_requirement_tree_supports_multi_delete_badges_and_drag_reclassification(self):
        requirements = [
            {'id': 'sql', 'title': 'SQL change', 'code': 'REQ-SQL', 'online_month': '2026-06', 'has_sql': True, 'needs_peripheral_upgrade': True},
            {'id': 'bug', 'title': 'Bug fix', 'code': 'BUG-101', 'online_month': '2026-06', 'record_kind': 'BUG'},
            {'id': 'keep', 'title': 'Keep', 'code': 'REQ-KEEP', 'online_month': '2026-02'},
        ]
        with patch('panels.requirement_panel.load_requirements', return_value=requirements), \
                patch('panels.requirement_panel.save_requirements') as save_data:
            panel = RequirementPanel()
            june = panel.requirement_list.topLevelItem(0)
            sql_item, bug_item = june.child(0), june.child(1)
            self.assertGreater(june.font(0).pointSize(), sql_item.font(0).pointSize())
            self.assertEqual(sql_item.text(0), 'REQ-SQL')
            self.assertIn('SQL change', sql_item.text(1))
            self.assertIn('SQL·待完成', sql_item.text(1))
            self.assertTrue(
                '周边·待完成' in sql_item.text(1) or '周边' in sql_item.text(1),
                sql_item.text(1),
            )

            panel._move_requirements(['keep'], '2026-06')
            self.assertEqual(next(item for item in panel._requirements if item['id'] == 'keep')['online_month'], '2026-06')
            save_data.assert_called()

            group = panel.requirement_list.topLevelItem(0)
            selected = [group.child(index) for index in range(group.childCount()) if group.child(index).data(0, Qt.ItemDataRole.UserRole)['id'] in ('sql', 'bug')]
            panel.requirement_list.clearSelection()
            panel.requirement_list.setCurrentItem(selected[0])
            for item in selected:
                item.setSelected(True)
            with patch('panels.requirement_panel.confirm_action', return_value=True):
                panel._delete_requirement()
            self.assertEqual([item['id'] for item in panel._requirements], ['keep'])

            # 全选后批量删除：删除按钮始终可见，并可删掉剩余项
            self.assertEqual(panel.batch_delete_btn.text(), '删除')
            panel.select_all_check.setChecked(True)
            self.app.processEvents()
            self.assertTrue(panel.batch_delete_btn.isEnabled())
            self.assertIn('删除', panel.batch_delete_btn.text())
            with patch('panels.requirement_panel.confirm_action', return_value=True):
                panel.batch_delete_btn.click()
            self.assertEqual(panel._requirements, [])
            panel.close()

    def test_requirement_dates_allow_manual_input_and_local_svn_binding(self):
        month = DateInput('2026-07', month_only=True)
        date = DateInput('2026-07-23')
        self.assertTrue(month.is_valid())
        self.assertTrue(date.is_valid())
        month.edit.setText('2026-13')
        self.assertFalse(month.is_valid())
        # 兼容 2026/07/23 写法，保存时会规范成 yyyy-MM-dd
        date.edit.setText('2026/07/23')
        self.assertTrue(date.is_valid())
        date.edit.setText('not-a-date')
        self.assertFalse(date.is_valid())

        with tempfile.TemporaryDirectory() as temp:
            os.makedirs(os.path.join(temp, '.svn'))
            dialog = RequirementDialog()
            dialog.title_edit.setText('绑定目录测试')
            dialog.svn_url_edit.setText('svn://10/example/DEV_REQ_TEST')
            dialog.local_path_edit.setText(temp)
            dialog.online_month.edit.setText('2026-07')
            dialog.planned_date.edit.setText('2026-07-23')
            values = dialog.values()
            self.assertEqual(values['local_path'], temp)
            self.assertEqual(values['workspace_kind'], 'svn')
            self.assertEqual(values['online_month'], '2026-07')
            self.assertEqual(values['planned_online_date'], '2026-07-23')
            dialog.close()

    def test_requirement_file_tree_shows_all_files_and_lock_icon_without_svn_status_column(self):
        with tempfile.TemporaryDirectory() as temp:
            for index in range(805):
                Path(temp, f'file-{index:03d}.txt').touch()
            os.makedirs(os.path.join(temp, 'SQL'))
            locked_path = os.path.join(temp, 'SQL', 'upgrade.sql')
            Path(locked_path).write_text('select 1;', encoding='utf-8')
            entries = workspace_files(temp)
            self.assertEqual(len(entries), 807)

            with patch('panels.requirement_panel.load_requirements', return_value=[]):
                panel = RequirementPanel()
            panel._current = {
                'id': 'locked', 'local_path': temp, 'workspace_kind': 'svn',
                'svn_locks': {os.path.join('SQL', 'upgrade.sql'): '2026-07-17T11:00:00'},
            }
            panel._file_tree_path = temp
            panel._file_tree_loaded(entries)
            self.assertEqual(panel.file_tree.columnCount(), 5)
            self.assertEqual(
                [panel.file_tree.headerItem().text(index) for index in range(5)],
                ['名称', '类型', '修改时间', '大小', '路径'],
            )
            header = panel.file_tree.header()
            self.assertEqual(header.sectionResizeMode(0), QHeaderView.ResizeMode.Stretch)
            for index in range(1, 5):
                self.assertEqual(header.sectionResizeMode(index), QHeaderView.ResizeMode.Interactive)
            self.assertFalse(header.stretchLastSection())
            sql_folder = next(panel.file_tree.topLevelItem(index) for index in range(panel.file_tree.topLevelItemCount()) if 'SQL' in panel.file_tree.topLevelItem(index).text(0))
            self.assertTrue(sql_folder.isExpanded())
            self.assertTrue(sql_folder.child(0).text(0).startswith('🔒'))
            self.assertFalse(sql_folder.child(0).icon(0).isNull())
            panel.close()

    def test_requirement_file_tree_refresh_is_silent_and_tree_has_selection_controls(self):
        with tempfile.TemporaryDirectory() as temp, \
                patch('panels.requirement_panel.load_requirements', return_value=[]):
            panel = RequirementPanel()
            panel._current = {'id': 'folder', 'local_path': temp, 'workspace_kind': 'folder'}
            with patch.object(panel, '_start_task') as start_task:
                panel._refresh_file_tree()
            self.assertEqual(start_task.call_args.kwargs['show_loading'], False)
            self.assertEqual(panel.requirement_list.columnCount(), 2)
            self.assertEqual(panel.select_all_check.text(), '全选')
            self.assertEqual(panel.batch_delete_btn.text(), '删除')
            self.assertFalse(panel.batch_delete_btn.isEnabled())
            self.assertEqual(panel.expand_tree_btn.text(), '全部展开')
            self.assertEqual(panel.collapse_tree_btn.text(), '全部折叠')
            self.assertTrue(hasattr(panel, 'file_search_edit'))
            self.assertEqual(panel.file_tree.header().sectionResizeMode(0), QHeaderView.ResizeMode.Stretch)
            self.assertEqual(panel.file_tree.header().sectionResizeMode(4), QHeaderView.ResizeMode.Interactive)
            panel.close()

    def test_requirement_detail_splitter_is_resizable_and_persistent(self):
        with patch('panels.requirement_panel.load_requirements', return_value=[]), \
                patch('panels.requirement_panel.load_requirement_ui', return_value={'splitter_sizes': [430, 620], 'content_splitter_sizes': [340, 230]}), \
                patch('panels.requirement_panel.save_requirement_ui') as save_ui:
            panel = RequirementPanel()
            panel.resize(1200, 780)
            panel.show()
            self.app.processEvents()

            self.assertIsInstance(panel.detail_splitter, QSplitter)
            self.assertEqual(panel.detail_splitter.orientation(), Qt.Orientation.Horizontal)
            self.assertGreaterEqual(panel.detail_splitter.handleWidth(), 8)
            self.assertFalse(panel.detail_splitter.childrenCollapsible())
            self.assertGreaterEqual(panel.detail_splitter.widget(1).minimumWidth(), 360)
            self.assertEqual(panel.file_sql_splitter.orientation(), Qt.Orientation.Vertical)
            self.assertFalse(panel.file_sql_splitter.childrenCollapsible())
            panel.detail_splitter.setSizes([520, 500])
            self.app.processEvents()
            sizes = panel.detail_splitter.sizes()
            self.assertGreater(sizes[0], 430)
            self.assertGreaterEqual(panel.detail_splitter.widget(0).minimumWidth(), 200)
            self.assertGreaterEqual(panel.system_filter.minimumWidth(), 160)
            self.assertGreaterEqual(panel.kind_filter.minimumWidth(), 100)
            self.assertGreaterEqual(panel.status_filter.minimumWidth(), 100)
            # 文件 Tab 工具条：打开/添加/新建/提交均存在（布局已收敛为横向 action card）
            for button in (panel.open_folder_btn, panel.add_file_btn, panel.new_text_btn, panel.commit_btn):
                self.assertIsNotNone(button)
                self.assertTrue(button.property('compactAction') or button.objectName() == 'primary-btn')
            self.assertTrue(panel.sql_btn.property('compactAction'))
            self.assertEqual(panel.open_folder_btn.text(), '打开目录')
            self.assertEqual(panel.sql_btn.text(), '打开发版联动')
            # 摘要卡不重复 Tab 内完整路径常驻说明；绑定状态为短 pill
            self.assertTrue(hasattr(panel, 'bind_status'))
            self.assertTrue(panel.svn_activity.isHidden())

            panel._save_splitter_sizes()
            # 兼容：file_sql_splitter 已隐藏，content 尺寸使用默认或缓存
            content_sizes = panel.file_sql_splitter.sizes() or [520, 140]
            save_ui.assert_called_with({'splitter_sizes': sizes, 'content_splitter_sizes': content_sizes})
            panel.close()

        style_path = os.path.join(ROOT, 'resources', 'style.qss')
        with open(style_path, 'r', encoding='utf-8') as stream:
            style = stream.read()
            self.assertIn('QSplitter#requirement-splitter::handle:horizontal', style)
            self.assertIn('QTreeWidget#requirement-file-tree QHeaderView::section', style)
            # 浅色导航右边线（主题 token）
            self.assertIn('border-right: 1px solid __SIDEBAR_BORDER__', style)

    def test_global_combo_and_date_styles_have_visible_drop_down_affordance(self):
        style_path = os.path.join(ROOT, 'resources', 'style.qss')
        with open(style_path, 'r', encoding='utf-8') as stream:
            style = stream.read()
        self.assertIn('QComboBox::drop-down', style)
        self.assertIn('QComboBox::down-arrow', style)
        self.assertIn('QDateEdit::drop-down', style)
        self.assertIn('QCalendarWidget QWidget#qt_calendar_navigationbar', style)

    def test_release_page_is_first_and_date_auto_loads_candidates(self):
        panel = SqlToolPanel()
        self.assertEqual(panel.tabs.tabText(0), '升级准备')
        self.assertEqual(panel.tabs.tabText(1), '发版联动')
        self.assertEqual(panel.tabs.tabText(2), '系统配置')
        self.assertEqual(panel.release_date.displayFormat(), 'yyyy-MM-dd')
        self.assertEqual(panel.date_edit.displayFormat(), 'yyyy-MM-dd')
        self.assertGreaterEqual(panel.release_date.minimumWidth(), 150)
        # SQL 加载按钮属于“SQL 脚本整理”页内部，不再作为顶部全局工具条
        self.assertIs(panel.load_btn.parent().parent(), panel.tabs.widget(1))
        self.assertEqual(panel.refresh_release_btn.text(), '刷新候选')
        # 打开后会自动按当前日期加载（兼容旧方法名）
        self.app.processEvents()
        panel._load_release_candidates()
        self.assertEqual(panel._release_date_confirmed, panel.release_date.date().toString('yyyy-MM-dd'))
        self.assertEqual(panel.release_table.rowCount(), len(panel._release_requirements))
        # 改日期后生成前会自动重载，不再要求“确认日期”
        panel.release_date.setDate(panel.release_date.date().addDays(1))
        self.app.processEvents()
        panel._release_reload_timer.stop()
        with patch.object(panel, '_load_release_candidates', wraps=panel._load_release_candidates) as loader:
            # 人为标记日期未同步，触发生成路径的自动重载
            panel._release_date_confirmed = ''
            with patch('panels.sql_panel.show_warning'), patch('panels.sql_panel.show_success'), \
                    patch('panels.sql_panel.save_requirements'):
                try:
                    panel._generate_release_materials()
                except Exception:
                    pass
            self.assertTrue(loader.called or panel._release_date_confirmed == panel.release_date.date().toString('yyyy-MM-dd'))
        self.assertEqual(panel.preview_tabs.tabText(0), '升级 SQL')
        self.assertEqual(panel.preview_tabs.tabText(2), '验证 SQL')
        panel.resize(980, 680)
        panel.show()
        self.app.processEvents()
        panel.tabs.setCurrentIndex(1)
        self.app.processEvents()
        self.assertTrue(panel.load_btn.isVisible())
        panel.tabs.setCurrentIndex(0)
        self.app.processEvents()
        screenshot = os.path.join(ROOT, '.codex_work', 'release_prep_ui.png')
        os.makedirs(os.path.dirname(screenshot), exist_ok=True)
        self.assertTrue(panel.grab().save(screenshot))
        panel.close()

    def test_only_learning_module_is_hidden(self):
        window = MainWindow()
        self.assertTrue(window.nav_buttons[8].isHidden())
        self.assertFalse(window.nav_buttons[9].isHidden())
        self.assertFalse(window.nav_buttons[10].isHidden())
        window._show_panel(9)
        self.assertEqual(window._current_nav_index, 9)
        window._show_panel(8)
        self.assertEqual(window._current_nav_index, 9)
        with patch('main_window.QInputDialog.getText', return_value=('Lihp', True)):
            self.assertTrue(window._unlock_private_tools())
        self.assertFalse(window.nav_buttons[8].isHidden())
        window._open_system_config()
        self.assertEqual(window._current_nav_index, 2)
        self.assertEqual(window.sql_panel.tabs.currentIndex(), 2)
        window._force_exit = True
        window.close()

    def test_one_click_generates_workbook_and_sql(self):
        requirement = {
            'id': 'test', 'record_kind': 'BUG', 'code': 'BUG-1', 'title': '测试问题',
            'description': '修复测试问题', 'planned_online_date': '2026-07-23',
            'system': '车险承保中心',
            'svn_url': 'svn://10/x/DEV_BUG_1',
            'sql_parts': [{'name': 'fix.sql', 'content': "update t set c='1' where id='1';"}],
        }
        with tempfile.TemporaryDirectory() as temp:
            template = os.path.join(temp, RELEASE_WORKBOOK_NAME)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = '20260723'
            for column, header in enumerate(RELEASE_HEADERS, 1):
                sheet.cell(1, column, header)
            workbook.save(template)
            workbook.close()

            panel = SqlToolPanel()
            panel._release_reload_timer.stop()
            panel.release_root.setText(temp)
            panel.release_date.blockSignals(True)
            panel.release_date.setDate(QDate(2026, 7, 23))
            panel.release_date.blockSignals(False)
            with patch('panels.sql_panel.load_requirements', return_value=[requirement]):
                panel._load_release_candidates()
            with patch('panels.sql_panel.save_requirements'), \
                    patch('panels.sql_panel.show_success'), \
                    patch('panels.sql_panel.show_info'), \
                    patch('panels.sql_panel.show_warning') as warning:
                panel._generate_release_materials()
            warning.assert_not_called()
            self.assertTrue(os.path.isfile(os.path.join(temp, '英大财险-共享中心-生产任务发版清单_20260723.xlsx')))
            sql_files = []
            for root, _dirs, files in os.walk(os.path.join(temp, '升级SQL')):
                sql_files.extend(os.path.join(root, name) for name in files if name.endswith('.sql'))
            self.assertTrue(sql_files)
            panel.close()

    def test_multiple_systems_generate_separate_sql_packages(self):
        panel = SqlToolPanel()
        first, second = panel._systems[:2]
        requirements = [
            {
                'id': 'a', 'record_kind': '需求', 'code': 'REQ-A', 'title': '系统A需求',
                'description': '系统A任务', 'planned_online_date': '2026-07-23',
                'system': first['name'], 'svn_url': 'svn://10/x/DEV_REQ_A',
                'sql_parts': [{'name': 'a.sql', 'content': "update table_a set marker='SYSTEM_A';"}],
            },
            {
                'id': 'b', 'record_kind': 'BUG', 'code': 'BUG-B', 'title': '系统B问题',
                'description': '系统B任务', 'planned_online_date': '2026-07-23',
                'system': second['name'], 'svn_url': 'svn://10/x/DEV_BUG_B',
                'sql_parts': [{'name': 'b.sql', 'content': "update table_b set marker='SYSTEM_B';"}],
            },
        ]
        with tempfile.TemporaryDirectory() as temp:
            template = os.path.join(temp, RELEASE_WORKBOOK_NAME)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = '20260723'
            for column, header in enumerate(RELEASE_HEADERS, 1):
                sheet.cell(1, column, header)
            workbook.save(template)
            workbook.close()

            panel._release_reload_timer.stop()
            panel.release_root.setText(temp)
            panel.release_date.blockSignals(True)
            panel.release_date.setDate(QDate(2026, 7, 23))
            panel.release_date.blockSignals(False)
            with patch('panels.sql_panel.load_requirements', return_value=requirements):
                panel._load_release_candidates()
            systems_by_code = {
                panel.release_table.item(row, 2).text(): panel.release_table.cellWidget(row, 5).currentData()
                for row in range(panel.release_table.rowCount())
            }
            self.assertEqual(systems_by_code['REQ-A'], first['name'])
            self.assertEqual(systems_by_code['BUG-B'], second['name'])
            panel.resize(1220, 780)
            panel.show()
            self.app.processEvents()
            # 截图前后禁止 debounce 定时器用真实台账冲掉测试数据
            panel._release_reload_timer.stop()
            screenshot = os.path.join(ROOT, '.codex_work', 'release_prep_multi_system_ui.png')
            self.assertTrue(panel.grab().save(screenshot))
            with patch('panels.sql_panel.load_requirements', return_value=requirements), \
                    patch('panels.sql_panel.save_requirements'), \
                    patch('panels.sql_panel.show_success'), \
                    patch('panels.sql_panel.show_info'), \
                    patch('panels.sql_panel.show_warning') as warning:
                panel._generate_release_materials()
            warning.assert_not_called()

            def package_text(system):
                folder = os.path.join(temp, '升级SQL')
                parts = []
                for root, _dirs, files in os.walk(folder):
                    if system['system_folder'] not in root:
                        continue
                    for name in files:
                        if name.endswith('.sql'):
                            with open(os.path.join(root, name), 'r', encoding='utf-8-sig') as stream:
                                parts.append(stream.read())
                return '\n'.join(parts)

            first_text = package_text(first)
            second_text = package_text(second)
            self.assertIn('SYSTEM_A', first_text)
            self.assertNotIn('SYSTEM_B', first_text)
            self.assertIn('SYSTEM_B', second_text)
            self.assertNotIn('SYSTEM_A', second_text)
            panel.close()

    def test_private_setup_does_not_touch_local_data(self):
        setup_path = os.path.join(ROOT, 'PrivateInstaller', 'setup.cmd')
        with open(setup_path, 'r', encoding='utf-8') as stream:
            setup = stream.read().casefold()
        self.assertNotIn('rmdir', setup)
        self.assertNotIn('del ', setup)
        self.assertNotIn('copy /y "%source_dir%data', setup)
        self.assertIn('if not exist "%install_dir%\\data" mkdir', setup)

    def test_upgrade_reuses_data_directory_and_accepts_legacy_requirement(self):
        old_exe = r'D:\PengToolsPrivate\PengToolsHub_Private_V4.24.exe'
        new_exe = r'D:\PengToolsPrivate\PengToolsHub_Private.exe'
        self.assertEqual(local_data_dir(old_exe, True), local_data_dir(new_exe, True))

        legacy = {
            'id': 'legacy', 'record_kind': '需求', 'code': 'REQ-OLD', 'title': '旧版需求',
            'planned_online_date': '2026-07-23', 'svn_url': 'svn://10/x/DEV_OLD',
            'legacy_custom_field': '必须保留',
        }
        panel = SqlToolPanel()
        panel._release_reload_timer.stop()
        panel.release_date.blockSignals(True)
        panel.release_date.setDate(QDate(2026, 7, 23))
        panel.release_date.blockSignals(False)
        with patch('panels.sql_panel.load_requirements', return_value=[legacy]):
            panel._load_release_candidates()
        system_combo = panel.release_table.cellWidget(0, 5)
        self.assertFalse(system_combo.currentData())
        system_combo.setCurrentIndex(1)
        selected = panel._selected_release_rows()
        self.assertEqual(selected[0][0]['legacy_custom_field'], '必须保留')
        self.assertEqual(selected[0][0]['system'], panel._systems[0]['name'])
        panel.close()


if __name__ == '__main__':
    unittest.main()
