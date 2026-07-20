# -*- coding: utf-8 -*-
import copy
import datetime
import os
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from tools.release_prep import (
    RELEASE_HEADERS, branch_name_from_svn, rank_requirements,
    release_row_from_requirement, update_release_workbook,
)


class ReleasePrepTests(unittest.TestCase):
    def _template(self, path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '20260723'
        for column, header in enumerate(RELEASE_HEADERS, 1):
            cell = sheet.cell(1, column, header)
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            sheet.column_dimensions[cell.column_letter].width = 18 + column
        sheet.cell(2, 1, '已有系统')
        sheet.cell(2, 2, '已有分支')
        for column in range(1, len(RELEASE_HEADERS) + 1):
            sheet.cell(3, column).fill = PatternFill('solid', fgColor='DDEBF7')
        workbook.save(path)
        workbook.close()

    def test_branch_and_date_priority(self):
        self.assertEqual(
            branch_name_from_svn('svn://10/x/DEV_prpcar_20260715-REQ-20260528-0006-SP/'),
            'DEV_prpcar_20260715-REQ-20260528-0006-SP',
        )
        records = [
            {'code': 'C', 'planned_online_date': ''},
            {'code': 'B', 'planned_online_date': '2026-07-10'},
            {'code': 'A', 'planned_online_date': '2026-07-23'},
            {'code': 'D', 'planned_online_date': '2026-08-01'},
        ]
        self.assertEqual([item['code'] for item in rank_requirements(records, '2026-07-23')], ['A', 'B', 'D', 'C'])

    def test_row_mapping(self):
        requirement = {
            'code': 'REQ-1', 'title': '测试需求', 'description': '任务说明',
            'owner': '张三', 'svn_url': 'svn://10/x/DEV_REQ_1',
            'needs_peripheral_upgrade': True, 'needs_interface_update': True,
        }
        row = release_row_from_requirement(requirement, '20260723')
        self.assertEqual(row['分支名称'], 'DEV_REQ_1')
        self.assertEqual(row['任务编号'], 'REQ-1')
        self.assertEqual(row['计划升级日期'], datetime.date(2026, 7, 23))
        self.assertEqual(row['平台/三方交互'], '是')
        self.assertEqual(row['相关文档'], '是')

    def test_append_and_create_sheet_without_touching_source(self):
        with tempfile.TemporaryDirectory() as temp:
            source = os.path.join(temp, 'source.xlsx')
            output_existing = os.path.join(temp, 'existing.xlsx')
            output_new = os.path.join(temp, 'new.xlsx')
            self._template(source)
            with open(source, 'rb') as stream:
                original = stream.read()
            row = release_row_from_requirement(
                {'code': 'REQ-1', 'title': '测试', 'svn_url': 'svn://10/x/DEV_REQ_1'},
                '20260723',
            )
            result = update_release_workbook(source, output_existing, '20260723', [row])
            self.assertEqual(result['start_row'], 3)
            existing = load_workbook(output_existing)
            self.assertEqual(existing['20260723'].cell(3, 2).value, 'DEV_REQ_1')
            existing.close()

            result = update_release_workbook(source, output_new, '20260724', [copy.deepcopy(row)])
            self.assertEqual(result['start_row'], 2)
            created = load_workbook(output_new)
            sheet = created['20260724']
            self.assertEqual(tuple(sheet.cell(1, column).value for column in range(1, 24)), RELEASE_HEADERS)
            self.assertEqual(sheet.column_dimensions['A'].width, 19)
            self.assertEqual(sheet.cell(2, 2).value, 'DEV_REQ_1')
            created.close()
            with open(source, 'rb') as stream:
                self.assertEqual(stream.read(), original)


if __name__ == '__main__':
    unittest.main()
